import os
import tkinter as tk
from tkinter import Toplevel, filedialog, messagebox
from tkinter.simpledialog import askinteger
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import tksheet

filenamesht = ''
sheetno = 0
data = ''
def load_xlsx():
    global filenamesht
    global sheetno
    global data
    filedir = filedialog.askopenfilename()
    filenamesht = filedir
    print(filenamesht)
    rows = []
    wb = load_workbook(filename = filedir)
    data = wb
    sheetsno = (len(data.sheetnames))
    if sheetsno == 1:
        sheetadd = data.sheetnames[0]
        sheetno = 0
        worksheet = data[sheetadd]
        print(worksheet.rows)
        for row in worksheet.iter_rows():
            xyz = []
            for cell in row:
                if cell.value == None:
                    xyz.append('')
                else:
                    xyz.append(cell.value)
            rows.append(xyz)
        sheet.set_sheet_data(rows)
    else:
        displaymsg = []
        for i in range (sheetsno):
            msg = str(i) + ' ' + data.sheetnames[i] + '\n'
            displaymsg.append(msg)
        strmsg = ' '.join(map(str, displaymsg))
        x = askinteger('Type the sheet number', strmsg )
        sheetno = x
        sheetadd = data.sheetnames[x]
        worksheet = data[sheetadd]
        print(worksheet.rows)
        for row in worksheet.iter_rows():
            xyz = []
            for cell in row:
                if cell.value == None:
                    xyz.append('')
                else:
                    xyz.append(cell.value)
            rows.append(xyz)
        sheet.set_sheet_data(rows)

    print(data.sheetnames[0])
    filename = str(filenamesht.split('/')[-1].split('.')[-2] + ' : ' + data.sheetnames[sheetno])
    uno_text.configure(text = (str(filename)))
    
def sort():
    global filenamesht
    global sheetno
    global data
    rows = []
    r_c = []
    sheet_choose = data.sheetnames[sheetno]
    print(data.sheetnames[0])
    worksheet = data[sheet_choose]
    
    comments_avail = []
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.comment:
                comments_avail.append(cell.comment.text)
    if len(comments_avail) > 1:
        simple_list = []
        [simple_list.append(x) for x in comments_avail if x not in simple_list]
        simple_list = sorted(simple_list)

        displaymsg = []
        for i in range (len(simple_list)):
            displaymsg.append([str(i) , simple_list[i], ''])

        comments_show = Toplevel()
        comments_show.geometry('400x800')
        frame_1 = tk.LabelFrame(comments_show, background= '#333333', foreground='#b4b4b4',  relief='flat')
        frame_1.pack(fill='x')
        frame_2 = tk.LabelFrame(comments_show, background= '#333333', foreground='#b4b4b4',  relief='flat')
        frame_2.pack(expand='True', fill='both')
        text_frame = tk.Label(frame_1, background='#333333', foreground='#b4b4b4', text='Select Comments for Sorting', font=('Arial', 15, 'bold'))
        text_frame.pack()
        sheet_show = tksheet.Sheet(frame_2)
        sheet_show.pack(fill='both', expand='True')
        sheet_show.change_theme(theme = "dark blue")
        displaymsg.pop(0)
        sheet_show.enable_bindings('all')
        for num in range(len(displaymsg)):
            sheet_show.create_checkbox(r=num,
                            c=2,
                            checked = False,
                            state = "normal",
                            redraw = True,
                            check_function = lambda x=None: p 
                            )
        sheet_show.set_sheet_data(displaymsg)

        def get_comment_value():
            selected_comments = []
            #list_of_row_no = sheet_show.get_selected_rows()
            #for i in list_of_row_no:
            #    row_data = sheet_show.get_row_data(r=i, return_copy = True)
            #    selected_comments.append(row_data[1])
            for data in displaymsg:
                if data[2] == True:
                    selected_comments.append(data[1])
            print(selected_comments)
            if len(selected_comments) > 0:
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.comment:
                            comments_avail.append(cell.comment.text)
                            #print(comment_ask,cell.comment.text)
                            if cell.comment.text in selected_comments:
                                rows.append(coordinate_from_string(cell.coordinate))
                for ig in rows:
                    zxx=worksheet.cell(row=ig[1], column=column_index_from_string(ig[0])).value
                    zxy = worksheet.cell(row=ig[1], column=1).value
                    r_c.append([zxy,zxx])
                sheet.set_sheet_data(r_c)
                comments_show.destroy()
            else:
                messagebox.showerror('Information', 'No row selected')
        comments_selection = tk.Button(comments_show, text='Sort using selected rows', command=lambda: get_comment_value())
        comments_selection.pack()
    else:
        messagebox.showerror('No comments', 'No comments found in sheet')

def change_sheet_tab():
    global sheetno
    global data
    rows = []
    sheetsno = (len(data.sheetnames))
    if sheetsno == 1:
        messagebox.showerror('Single Tab','Only single tab in sheet')
        return
        sheetadd = data.sheetnames[0]
        worksheet = data[sheetadd]
        print(worksheet.rows)
        for row in worksheet.iter_rows():
            xyz = []
            for cell in row:
                if cell.value == None:
                    xyz.append('')
                else:
                    xyz.append(cell.value)
            rows.append(xyz)
        sheet.set_sheet_data(rows)
    else:
        displaymsg = []
        for i in range (sheetsno):
            msg = str(i) + ' ' + data.sheetnames[i] + '\n'
            displaymsg.append(msg)
        strmsg = ' '.join(map(str, displaymsg))
        x = askinteger('Type the sheet number', strmsg )
        sheetno = int(x)
        sheetadd = data.sheetnames[x]
        worksheet = data[sheetadd]
        print(worksheet.rows)
        for row in worksheet.iter_rows():
            xyz = []
            for cell in row:
                if cell.value == None:
                    xyz.append('')
                else:
                    xyz.append(cell.value)
            rows.append(xyz)
        sheet.set_sheet_data(rows)
    filename = filenamesht.split('/')[-1].split('.')[-2] + ' : ' + data.sheetnames[sheetno]
    uno_text.configure(text = (str(filename)))


def p ():
    pass

root = tk.Tk()
root.geometry('800x800')
root.title('XLSort')
root.configure(bg='#333333')
frame_uno = tk.LabelFrame(root, background= '#333333', foreground='#b4b4b4',  relief='flat')
frame_uno.pack(fill='x')
uno_text = tk.Label(frame_uno, background='#333333', foreground='#b4b4b4', text='', font=('Arial', 15, 'bold'))
uno_text.pack()
frame_dos = tk.LabelFrame(root, background= '#333333', foreground='#b4b4b4',  relief='flat')
frame_dos.pack(fill='both', expand='True')
frame_tres = tk.LabelFrame(root, background= '#333333', foreground='#b4b4b4',  relief='flat')
frame_tres.pack(fill='x', pady=(5,10))
sheet = tksheet.Sheet(frame_dos)
sheet.pack(fill='both', expand='True')
sheet.change_theme(theme = "dark blue")
loadbutton = tk.Button(frame_tres, text='Load xlsx', command=lambda: load_xlsx())
loadbutton.pack(side='left',padx=5)
sortbutton = tk.Button(frame_tres, text='Sort using comments', command=lambda: sort())
sortbutton.pack(side='left',padx=5)
change_tab = tk.Button(frame_tres, text='Change tab', command=lambda: change_sheet_tab())
change_tab.pack(side='left',padx=5)
sheet.enable_bindings(("all"))
root.mainloop()

